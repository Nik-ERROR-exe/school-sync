from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.database import get_db
from app.api.deps import require_admin
from app.schemas.result import ResultResponse, ResultApproval
from app.services.result_service import get_results_by_status, approve_result, update_result
from app.models.teacher import Teacher
from app.models.result import Result
from app.models.student import Student
from app.models.school_class import SchoolClass
from app.models.subject import Subject
from app.models.exam_type import ExamType

router = APIRouter(
    prefix="/admin/results",
    tags=["Admin - Student Results"],
    dependencies=[Depends(require_admin)]
)


def calculate_grade(percentage: float) -> str:
    """Calculate grade based on percentage."""
    if percentage >= 90: return "A+"
    elif percentage >= 80: return "A"
    elif percentage >= 70: return "B"
    elif percentage >= 60: return "C"
    elif percentage >= 50: return "D"
    elif percentage >= 40: return "E"
    else: return "F"


@router.get("/", response_model=List[ResultResponse])
def list_results(
    class_id: Optional[int] = Query(None, description="Filter by class ID"),
    exam_type_id: Optional[int] = Query(None, description="Filter by exam type ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
    db: Session = Depends(get_db)
):
    """Retrieves all student results. Can filter by class, exam type, and status."""
    query = db.query(Result)
    
    if class_id:
        query = query.filter(Result.class_id == class_id)
    if exam_type_id:
        query = query.filter(Result.exam_type_id == exam_type_id)
    if status:
        query = query.filter(Result.status == status)
    
    results = query.all()
    
    response_data = []
    for r in results:
        # Get related data using direct queries (no relationships)
        student = db.query(Student).filter(Student.id == r.student_id).first()
        subject = db.query(Subject).filter(Subject.id == r.subject_id).first()
        exam_type = db.query(ExamType).filter(ExamType.id == r.exam_type_id).first()
        
        student_name = student.name if student else None
        student_roll_no = student.roll_no if student else None
        student_class = None
        student_division = None
        if student:
            school_class = db.query(SchoolClass).filter(SchoolClass.id == student.class_id).first()
            if school_class:
                student_class = school_class.class_name
                student_division = school_class.division
        
        response_data.append(
            ResultResponse(
                id=r.id,
                student_id=r.student_id,
                student_roll_no=student_roll_no,
                student_name=student_name,
                student_class=student_class,
                student_division=student_division,
                subject_id=r.subject_id,
                subject_name=subject.subject_name if subject else None,
                subject_code=subject.code if subject else None,
                exam_type_id=r.exam_type_id,
                exam_type_name=exam_type.name if exam_type else None,
                marks_obtained=r.marks_obtained,
                total_marks=r.total_marks,
                percentage=r.percentage,
                grade=r.grade,
                status=r.status,
                submitted_by_id=r.submitted_by_id,
                approved_by_id=r.approved_by_id
            )
        )
    return response_data


@router.get("/class/{class_id}/exam/{exam_type_id}")
def get_results_by_class_and_exam(
    class_id: int,
    exam_type_id: int,
    db: Session = Depends(get_db),
    admin: Teacher = Depends(require_admin)
):
    """Get all results for a specific class and exam type."""
    # Get all students in the class
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.roll_no).all()
    
    # Get all subjects for the class and exam
    subjects = db.query(Subject).join(
        Result, Result.subject_id == Subject.id
    ).filter(
        Result.class_id == class_id,
        Result.exam_type_id == exam_type_id
    ).distinct().order_by(Subject.subject_name).all()
    
    # Get all results
    results = db.query(Result).filter(
        Result.class_id == class_id,
        Result.exam_type_id == exam_type_id
    ).all()
    
    # Build response: students with their marks for each subject
    response_data = []
    for student in students:
        student_data = {
            "student_id": student.id,
            "roll_no": student.roll_no,
            "name": student.name,
            "subjects": []
        }
        
        for subject in subjects:
            result = next(
                (r for r in results if r.student_id == student.id and r.subject_id == subject.id),
                None
            )
            
            student_data["subjects"].append({
                "subject_id": subject.id,
                "subject_name": subject.subject_name,
                "marks_obtained": result.marks_obtained if result else None,
                "total_marks": result.total_marks if result else None,
                "percentage": result.percentage if result else None,
                "grade": result.grade if result else None,
                "status": result.status if result else None,
                "result_id": result.id if result else None
            })
        
        response_data.append(student_data)
    
    return {
        "class_id": class_id,
        "exam_type_id": exam_type_id,
        "students": response_data,
        "subjects": [{"id": s.id, "name": s.subject_name} for s in subjects]
    }


@router.put("/{id}/approve", response_model=ResultResponse)
def approve_student_result(
    id: int,
    approval: ResultApproval,
    current_admin: Teacher = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Approves or rejects a submitted student result record."""
    result = db.query(Result).filter(Result.id == id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")
    
    if approval.approved:
        result.status = "approved"
    else:
        result.status = "rejected"
    
    result.approved_by_id = current_admin.id
    result.approved_at = datetime.utcnow()
    
    db.commit()
    db.refresh(result)
    
    # Get related data using direct queries
    student = db.query(Student).filter(Student.id == result.student_id).first()
    subject = db.query(Subject).filter(Subject.id == result.subject_id).first()
    exam_type = db.query(ExamType).filter(ExamType.id == result.exam_type_id).first()
    
    student_name = student.name if student else None
    student_roll_no = student.roll_no if student else None
    student_class = None
    student_division = None
    if student:
        school_class = db.query(SchoolClass).filter(SchoolClass.id == student.class_id).first()
        if school_class:
            student_class = school_class.class_name
            student_division = school_class.division
    
    return ResultResponse(
        id=result.id,
        student_id=result.student_id,
        student_roll_no=student_roll_no,
        student_name=student_name,
        student_class=student_class,
        student_division=student_division,
        subject_id=result.subject_id,
        subject_name=subject.subject_name if subject else None,
        subject_code=subject.code if subject else None,
        exam_type_id=result.exam_type_id,
        exam_type_name=exam_type.name if exam_type else None,
        marks_obtained=result.marks_obtained,
        total_marks=result.total_marks,
        percentage=result.percentage,
        grade=result.grade,
        status=result.status,
        submitted_by_id=result.submitted_by_id,
        approved_by_id=result.approved_by_id
    )


@router.put("/{id}")
def update_result_endpoint(
    id: int,
    data: dict,
    db: Session = Depends(get_db),
    admin: Teacher = Depends(require_admin)
):
    """Update a result (for admin auto-save)."""
    marks_obtained = data.get("marks_obtained")
    total_marks = data.get("total_marks")
    
    if marks_obtained is None and total_marks is None:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = db.query(Result).filter(Result.id == id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")
    
    if marks_obtained is not None:
        result.marks_obtained = marks_obtained
    if total_marks is not None:
        result.total_marks = total_marks
    
    if result.total_marks > 0:
        result.percentage = (result.marks_obtained / result.total_marks) * 100
        p = result.percentage
        if p >= 90: result.grade = "A+"
        elif p >= 80: result.grade = "A"
        elif p >= 70: result.grade = "B"
        elif p >= 60: result.grade = "C"
        elif p >= 50: result.grade = "D"
        elif p >= 40: result.grade = "E"
        else: result.grade = "F"
    
    db.commit()
    db.refresh(result)
    return {"message": "Result updated successfully", "result": result}


@router.put("/{id}/reject")
def reject_result(
    id: int,
    db: Session = Depends(get_db),
    admin: Teacher = Depends(require_admin)
):
    """Reject a submitted result."""
    result = db.query(Result).filter(Result.id == id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")
    
    result.status = "rejected"
    result.approved_by_id = admin.id
    result.approved_at = datetime.utcnow()
    
    db.commit()
    db.refresh(result)
    return {"message": "Result rejected successfully"}


# ============================================================
# EXPORT TO EXCEL
# ============================================================
@router.get("/export")
def export_results_excel(
    class_id: int = Query(..., description="Class ID"),
    exam_type_id: int = Query(..., description="Exam Type ID"),
    db: Session = Depends(get_db),
    admin: Teacher = Depends(require_admin)
):
    """
    Export results to Excel file for a specific class and exam type.
    """
    # Get class info
    school_class = db.query(SchoolClass).filter(SchoolClass.id == class_id).first()
    if not school_class:
        raise HTTPException(status_code=404, detail="Class not found")
    
    # Get exam type info
    exam_type = db.query(ExamType).filter(ExamType.id == exam_type_id).first()
    if not exam_type:
        raise HTTPException(status_code=404, detail="Exam type not found")
    
    # Get students with their results
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.roll_no).all()
    
    # Get all subjects for this class and exam
    subjects = db.query(Subject).join(
        Result, Result.subject_id == Subject.id
    ).filter(
        Result.class_id == class_id,
        Result.exam_type_id == exam_type_id
    ).distinct().order_by(Subject.subject_name).all()
    
    # Get all results
    results = db.query(Result).filter(
        Result.class_id == class_id,
        Result.exam_type_id == exam_type_id
    ).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # --- HEADER STYLES ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- CREATE HEADER ROW ---
    headers = ["Roll No", "Student Name"]
    for subject in subjects:
        headers.append(subject.subject_name)
    headers.extend(["Total", "Percentage", "Grade"])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # --- FILL DATA ROWS ---
    row_num = 2
    for student in students:
        total_obtained = 0
        total_max = 0
        has_results = False
        
        row_data = [student.roll_no, student.name]
        
        for subject in subjects:
            result = next(
                (r for r in results if r.student_id == student.id and r.subject_id == subject.id),
                None
            )
            if result:
                row_data.append(result.marks_obtained)
                total_obtained += result.marks_obtained
                total_max += result.total_marks
                has_results = True
            else:
                row_data.append("-")
        
        if has_results and total_max > 0:
            percentage = (total_obtained / total_max) * 100
            grade = calculate_grade(percentage)
            row_data.append(f"{total_obtained}/{total_max}")
            row_data.append(round(percentage, 2))
            row_data.append(grade)
        else:
            row_data.append("-")
            row_data.append("-")
            row_data.append("-")
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style
        
        row_num += 1
    
    # --- AUTO-ADJUST COLUMN WIDTHS ---
    for col in range(1, len(headers) + 1):
        column_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
        max_length = 0
        for row in range(1, row_num):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 25)
    
    # --- GENERATE FILENAME ---
    filename = f"{school_class.class_name}{school_class.division}_{exam_type.name.replace(' ', '_')}.xlsx"
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )