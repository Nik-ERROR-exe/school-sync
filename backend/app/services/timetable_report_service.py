"""PDF / Excel export builders for the saved master timetable.

Both exporters consume the same per-class day x period grid produced by
`build_timetable_grids`, so the layout logic lives in one place. The styling
follows `report_service.py` (reportlab for PDF, openpyxl for Excel).
"""

import html
import json
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.date_utils import DAY_NAMES, format_time, int_to_day
from app.models.timetable import TimetableSlot
from app.models.timetable_settings import TimetableSettings


def _resolve_days(slots: List[TimetableSlot], settings: Optional[TimetableSettings]) -> List[str]:
    """Determine the day columns: from saved settings, else days present in slots."""
    if settings and settings.school_days:
        try:
            saved_days = json.loads(settings.school_days)
            if saved_days:
                return [d for d in DAY_NAMES if d in saved_days]
        except (ValueError, TypeError):
            pass

    present = {int_to_day(s.day_of_week) for s in slots}
    return [d for d in DAY_NAMES if d in present]


def build_timetable_grids(
    slots: List[TimetableSlot],
    settings: Optional[TimetableSettings],
) -> List[dict]:
    """Group enriched timetable slots into per-class day x period grids.

    Each returned grid has the shape::

        {
            "class_id": int,
            "class_name": str,
            "division": str,
            "days": [day_name, ...],
            "rows": {period_number: {day_name: cell_text}},
        }

    Cell text is ``"Subject\\nTeacher"`` for a scheduled slot, ``"LUNCH"`` for
    the lunch marker (``subject_id == 0`` saved by ``save_timetable``), or an
    empty string when no class runs that period.
    """
    days = _resolve_days(slots, settings)

    by_class: dict = {}
    for s in slots:
        by_class.setdefault(s.class_id, []).append(s)

    grids: List[dict] = []
    for class_id, class_slots in by_class.items():
        cls = class_slots[0].school_class
        class_name = cls.class_name if cls else str(class_id)
        division = cls.division if cls else ""
        max_period = max(s.period_number for s in class_slots)

        rows = {p: {day: "" for day in days} for p in range(1, max_period + 1)}
        for s in class_slots:
            day = int_to_day(s.day_of_week)
            if day not in days:
                continue
            if s.subject_id == 0:
                rows[s.period_number][day] = "LUNCH"
                continue
            subject = s.subject.subject_name if s.subject else f"Subject #{s.subject_id}"
            teacher = s.teacher.name if (s.teacher and s.teacher_id) else ""
            rows[s.period_number][day] = subject + (f"\n{teacher}" if teacher else "")

        grids.append({
            "class_id": class_id,
            "class_name": class_name,
            "division": division,
            "days": days,
            "rows": rows,
        })

    return grids


def _settings_line(settings: Optional[TimetableSettings], grids: List[dict]) -> str:
    """Short settings summary shown under the PDF title."""
    parts: List[str] = []
    if settings:
        if settings.start_time:
            parts.append(f"Start {format_time(settings.start_time)}")
        if settings.period_duration:
            parts.append(f"{settings.period_duration} min periods")
    if grids:
        parts.append(f"{len(grids[0]['days'])} school days")
    return "  |  ".join(parts)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_timetable_pdf(
    grids: List[dict],
    settings: Optional[TimetableSettings],
    school_name: str = "SchoolSync Academy",
) -> BytesIO:
    """Render each class grid as a styled reportlab table in a single PDF."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TimetableTitle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=24,
        textColor=colors.HexColor('#1A365D'), alignment=1, spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        'TimetableSubTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=12,
        textColor=colors.HexColor('#4A5568'), alignment=1, spaceAfter=6,
    )
    settings_style = ParagraphStyle(
        'TimetableSettings', parent=styles['Normal'],
        fontName='Helvetica', fontSize=9,
        textColor=colors.HexColor('#718096'), alignment=1, spaceAfter=24,
    )
    class_heading_style = ParagraphStyle(
        'ClassHeading', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=14,
        textColor=colors.HexColor('#1A365D'), spaceBefore=18, spaceAfter=8,
    )
    header_cell_style = ParagraphStyle(
        'HeaderCell', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=9, textColor=colors.whitesmoke,
    )
    cell_style = ParagraphStyle(
        'CellText', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#2D3748'),
    )

    story.append(Paragraph(school_name, title_style))
    story.append(Paragraph("MASTER TIMETABLE", subtitle_style))
    story.append(Paragraph(_settings_line(settings, grids), settings_style))

    # letter is 612pt wide; minus 80pt margins => 532pt printable area
    printable_width = 532

    for grid in grids:
        heading = f"Class {grid['class_name']}"
        if grid['division']:
            heading += f" - Division {grid['division']}"
        story.append(Paragraph(heading, class_heading_style))

        header_row = [Paragraph("<b>Period</b>", header_cell_style)]
        header_row += [Paragraph(f"<b>{day}</b>", header_cell_style) for day in grid['days']]

        data = [header_row]
        for period in sorted(grid['rows'].keys()):
            row = [Paragraph(str(period), cell_style)]
            for day in grid['days']:
                text = grid['rows'][period].get(day, "")
                if text:
                    cell_html = "<br/>".join(html.escape(part) for part in text.split("\n"))
                    row.append(Paragraph(cell_html, cell_style))
                else:
                    row.append(Paragraph("&nbsp;", cell_style))
            data.append(row)

        day_col = max((printable_width - 40) // max(len(grid['days']), 1), 60)
        table = Table(data, colWidths=[40] + [day_col] * len(grid['days']))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B6CB0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F7FAFC')),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
        ]))
        story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _excel_sheet_title(grid: dict) -> str:
    """Sheet title like 'Class 5A', sanitized for Excel's 31-char / illegal-char rules."""
    raw = f"Class {grid['class_name']}"
    if grid['division']:
        raw += f" - {grid['division']}"
    for ch in "[]:*?/\\":
        raw = raw.replace(ch, "-")
    return raw[:31]


def generate_timetable_excel(
    grids: List[dict],
    settings: Optional[TimetableSettings],
    school_name: str = "SchoolSync Academy",
) -> BytesIO:
    """Render one styled day x period sheet per class in an .xlsx workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet; we add one per class

    title_font = Font(name="Calibri", size=14, bold=True, color="1A365D")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="4A5568")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)

    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0"),
    )

    for grid in grids:
        ws = wb.create_sheet(title=_excel_sheet_title(grid))
        ws.views.sheetView[0].showGridLines = True

        ncols = 1 + len(grid['days'])
        last_col = get_column_letter(ncols)

        # Title / subtitle rows
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = school_name
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f"A2:{last_col}2")
        subtitle = f"Master Timetable - Class {grid['class_name']}"
        if grid['division']:
            subtitle += f" - Division {grid['division']}"
        ws["A2"] = subtitle
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = center_align
        ws.row_dimensions[2].height = 18

        ws.append([])  # spacer row 3

        # Header row (row 4)
        headers = ["Period"] + grid['days']
        ws.append(headers)
        header_row = 4
        ws.row_dimensions[header_row].height = 22
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for idx, period in enumerate(sorted(grid['rows'].keys())):
            row_data = [period]
            row_data += [grid['rows'][period].get(day, "") for day in grid['days']]
            ws.append(row_data)
            curr_row = header_row + 1 + idx
            ws.row_dimensions[curr_row].height = 24
            for col_idx in range(1, ncols + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_align if col_idx > 1 else center_align
                if idx % 2 == 1:
                    cell.fill = alt_row_fill

        # Auto-adjust column widths (skip merged title/subtitle rows)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.row in (1, 2):
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value).split("\n")[0]))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    if len(grids) == 0:
        ws = wb.create_sheet(title="Timetable")
        ws["A1"] = "No timetable data."

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
