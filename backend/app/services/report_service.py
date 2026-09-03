import os
from io import BytesIO
from datetime import datetime
from typing import List
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.result import Result
from app.config import settings
from app.services.result_service import calculate_overall_grade, get_grading_scale_group

def generate_results_pdf(results: List[Result], school_name: str = "SchoolSync Academy") -> BytesIO:
    """
    Generates a high-quality, professional PDF report of approved student results
    using ReportLab.
    """
    buffer = BytesIO()
    
    # Page setup
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    story = []
    
    styles = getSampleStyleSheet()
    
    # Premium Typography & Color styles matching modern designs
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1A365D'),  # Deep Navy Blue
        alignment=1,  # Center
        spaceAfter=8
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#4A5568'),  # Dark Gray
        alignment=1,
        spaceAfter=24
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#2D3748')
    )
    
    cell_style_bold = ParagraphStyle(
        'CellTextBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    
    # Document Header
    story.append(Paragraph(school_name, title_style))
    story.append(Paragraph("OFFICIAL STUDENT PERFORMANCE REPORT (APPROVED BATCH)", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Table headers and contents
    data = [[
        Paragraph("<b>Roll No</b>", cell_style_bold),
        Paragraph("<b>Student Name</b>", cell_style_bold),
        Paragraph("<b>Class</b>", cell_style_bold),
        Paragraph("<b>Subject</b>", cell_style_bold),
        Paragraph("<b>Exam Type</b>", cell_style_bold),
        Paragraph("<b>Marks</b>", cell_style_bold),
        Paragraph("<b>Grade</b>", cell_style_bold)
    ]]
    
    for r in results:
        student_name = r.student.name if r.student else "N/A"
        roll_no = r.student.roll_no if r.student else "N/A"
        class_name = f"{r.student.school_class.class_name}{r.student.school_class.division}" if r.student and r.student.school_class else "N/A"
        subject_name = r.subject.subject_name if r.subject else "N/A"
        exam_name = r.exam_type.name if r.exam_type else "N/A"
        marks_str = f"{r.marks_obtained} / {r.total_marks} ({r.percentage}%)"
        grade_str = r.grade
        
        data.append([
            Paragraph(roll_no, cell_style),
            Paragraph(student_name, cell_style),
            Paragraph(class_name, cell_style),
            Paragraph(subject_name, cell_style),
            Paragraph(exam_name, cell_style),
            Paragraph(marks_str, cell_style),
            Paragraph(grade_str, cell_style)
        ])
        
    # Table layouts - margins: letter is 612 wide. 612 - 80 margins = 532 printable area
    t = Table(data, colWidths=[65, 120, 50, 105, 75, 82, 35])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2B6CB0')),  # Soft Teal/Blue Header
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),  # Subtle border lines
        # Alternating row background colors
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F7FAFC')),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 6),
    ]))
    
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_results_excel(results: List[Result], school_name: str = "Amarkor Vidyalaya") -> BytesIO:
    """
    Generates a premium, formatted Microsoft Excel sheet containing the approved results.
    Header row shows Class, Academic Year, and Exam Type side-by-side (separate cells).
    The results table is pivoted: one row per student, with each subject as a column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Results"

    # Make sure gridlines are visible
    ws.views.sheetView[0].showGridLines = True

    # Styled Font Family
    title_font = Font(name="Calibri", size=16, bold=True, color="1A365D")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="4A5568")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    header_value_font = Font(name="Calibri", size=14, bold=True, color="1A365D")

    # Fill colors
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Borders
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )

    # ------------------------------------------------------------------
    # Derive Class name, Academic Year, and Exam Type from the results.
    # All results in a single export share the same class and exam type.
    # ------------------------------------------------------------------
    class_display = "N/A"
    exam_type_name = "N/A"
    if results and len(results) > 0:
        first = results[0]
        if first.student and first.student.school_class:
            sc = first.student.school_class
            class_display = f"{sc.class_name}{sc.division}"
        if first.exam_type:
            exam_type_name = first.exam_type.name

    # Derive academic year from ACADEMIC_TERM_START (e.g. "2026-04-01" -> "2026-27")
    try:
        term_start = datetime.strptime(settings.ACADEMIC_TERM_START, "%Y-%m-%d")
        start_year = term_start.year
        end_year = (start_year + 1) % 100
        academic_year = f"{start_year}-{end_year:02d}"
    except (ValueError, AttributeError):
        academic_year = "N/A"

    # ------------------------------------------------------------------
    # Collect subjects first so we can size the merged title row correctly
    # ------------------------------------------------------------------
    seen_subjects = []
    seen_set = set()
    for r in results:
        if r.subject:
            subj_name = r.subject.subject_name
            if subj_name not in seen_set:
                seen_set.add(subj_name)
                seen_subjects.append(subj_name)

    total_cols = 2 + len(seen_subjects) + 3  # Roll No + Name + subjects + Total + % + Grade

    # ------------------------------------------------------------------
    # Row 1 — School name (merged across all table columns, centered title)
    # ------------------------------------------------------------------
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = school_name
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    # ------------------------------------------------------------------
    # Row 2 — Class | Academic Year | Exam Type (side-by-side, separate cells)
    # ------------------------------------------------------------------
    ws["A2"] = class_display
    ws["B2"] = academic_year
    ws["C2"] = exam_type_name

    for col_idx in range(1, 4):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_value_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[2].height = 25

    # ------------------------------------------------------------------
    # Row 3 — Empty spacing row
    # ------------------------------------------------------------------
    ws.append([])

    # ------------------------------------------------------------------
    # Row 4 — Table header (pivoted: subjects become columns)
    # ------------------------------------------------------------------
    headers = ["Roll No", "Student Name"] + seen_subjects + ["Total Marks", "Percentage", "Grade"]
    header_row = 4
    ws.append(headers)

    ws.row_dimensions[header_row].height = 25

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ------------------------------------------------------------------
    # Pivoted data rows — one row per student
    # ------------------------------------------------------------------
    # Group results by student, preserving roll-number ordering
    students_map = {}
    student_order = []

    for r in results:
        if not r.student:
            continue
        sid = r.student.id
        if sid not in students_map:
            students_map[sid] = {
                "roll_no": r.student.roll_no,
                "name": r.student.name,
                "marks_by_subject": {},
            }
            student_order.append((sid, r.student.roll_no))
        if r.subject:
            students_map[sid]["marks_by_subject"][r.subject.subject_name] = r.marks_obtained

    # Sort students by numeric roll number
    def sort_key(item):
        roll = item[1]
        try:
            return (0, int(roll))
        except (ValueError, TypeError):
            return (1, roll if roll else "")
    student_order.sort(key=sort_key)

    # ------------------------------------------------------------------
    # Derive grading scale group from class name for overall grade calculation
    # ------------------------------------------------------------------
    scale_group = "STD_1_8"
    if results and len(results) > 0:
        first = results[0]
        if first.student and first.student.school_class:
            scale_group = get_grading_scale_group(first.student.school_class.class_name)

    for idx, (sid, _) in enumerate(student_order):
        row_data = students_map[sid]
        marks_map = row_data["marks_by_subject"]

        row_values = [
            row_data["roll_no"],
            row_data["name"],
        ]
        # Add marks for each subject column
        for subj in seen_subjects:
            row_values.append(marks_map.get(subj, ""))

        # Check if student has marks filled for ALL subjects
        has_all_marks = (
            len(seen_subjects) > 0 and
            all(marks_map.get(subj) is not None and marks_map.get(subj) != "" for subj in seen_subjects)
        )

        if has_all_marks:
            # Calculate overall totals per student (matches calculate_class_overall_results logic)
            student_result_objs = [
                r for r in results
                if r.student and r.student.id == sid and r.subject
            ]
            total_obtained = sum(float(r.marks_obtained) for r in student_result_objs)
            total_max = sum(float(r.total_marks) for r in student_result_objs)

            row_values.append(total_obtained)

            # Percentage and grade (consistent with calculate_class_overall_results)
            pct = round((total_obtained * 100.0) / total_max, 2) if total_max > 0 else 0.0
            grd = calculate_overall_grade(pct, scale_group)

            row_values.append(pct)
            row_values.append(grd)
        else:
            row_values.append("")
            row_values.append("")
            row_values.append("")

        ws.append(row_values)
        curr_row = header_row + 1 + idx
        ws.row_dimensions[curr_row].height = 20

        # Format columns styling
        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            if col_idx in [2]:  # Name
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            # Alternating rows coloring
            if idx % 2 == 1:
                cell.fill = alt_row_fill

            # Format percentage cell (second-to-last column)
            if col_idx == total_cols - 1:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'

            # Format numeric columns (not Roll No or Name)
            if col_idx not in [1, 2]:
                val = cell.value
                if isinstance(val, (int, float)):
                    if col_idx < total_cols - 2:  # Subject marks columns
                        cell.number_format = '0.0'
                    elif col_idx == total_cols - 2:  # Total Marks column
                        cell.number_format = '0.0'

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:  # Skip title, header, spacing row
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
