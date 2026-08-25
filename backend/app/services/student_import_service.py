import csv
import io
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

TEMPLATE_HEADERS = ["Roll No", "Name", "Class", "Division"]


class InvalidFileError(Exception):
    pass


def _canon(key: str) -> str:
    # "Roll No", "roll_no", "Roll Number" all -> "rollno"
    return re.sub(r"[\s_\-\.]+", "", (key or "").lower())


# Canonical field name -> accepted header aliases (in canonical form).
HEADER_ALIASES: Dict[str, set] = {
    "roll_no": {"rollno", "roll", "rollnumber", "enrollment", "admno"},
    "name": {"name", "studentname", "fullname"},
    "class_name": {"class", "classname", "standard", "std", "cls"},
    "division": {"division", "div", "section"},
}


def normalize_header_row(raw_headers: List[Optional[str]]) -> Dict[str, int]:
    """Canonical field name -> column index, resolved via aliases."""
    idx_map: Dict[str, int] = {}
    for idx, h in enumerate(raw_headers):
        if h is None:
            continue
        key = _canon(str(h))
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in idx_map:
                idx_map[field] = idx
                break
    return idx_map


def _row_has_data(vals) -> bool:
    return any(v is not None and str(v).strip() != "" for v in vals)


def parse_student_file(filename: str, content: bytes) -> Tuple[Dict[str, int], List[Dict]]:
    """Returns (header_idx_map, list of row dicts keyed by canonical field).

    Row dicts only contain keys for columns that exist in the file; a cell that
    is present-but-empty yields None, an absent column yields no key.
    """
    lower = (filename or "").lower()

    if lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            raw_headers = list(next(ws.iter_rows(values_only=True), None) or [])
            idx_map = normalize_header_row(raw_headers)
            rows = [
                {field: (vals[idx] if idx < len(vals) else None)
                 for field, idx in idx_map.items()}
                for vals in ws.iter_rows(values_only=True)
                if _row_has_data(vals)
            ]
        finally:
            wb.close()
        return idx_map, rows

    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")  # handles Excel BOM
        reader = csv.reader(io.StringIO(text))
        raw_headers = next(reader, None)
        if raw_headers is None:
            return {}, []
        idx_map = normalize_header_row(raw_headers)
        rows = [
            {field: (vals[idx] if idx < len(vals) else None)
             for field, idx in idx_map.items()}
            for vals in reader
            if _row_has_data(vals)
        ]
        return idx_map, rows

    raise InvalidFileError(
        "Unsupported file type. Only .xlsx and .csv are supported (legacy .xls is not)."
    )


def generate_student_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["1", "Example Student", "1", "A"])
    for col, width in {"A": 12, "B": 30, "C": 10, "D": 10}.items():
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
